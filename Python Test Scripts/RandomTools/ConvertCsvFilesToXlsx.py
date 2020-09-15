import os
import sys
import csv
import argparse
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

def convert_csv_files_to_xlsx(xlsx_location, csv_folder):
    # Set the maximum field size limit
    max_int = sys.maxsize
    while True:
        try:
            csv.field_size_limit(max_int)
            break
        except:
            max_int = int(max_int / 10)

    # Create the Excel workbook option
    wb = Workbook()

    # Delete the default sheet
    ws = wb.active
    wb.remove(ws)

    # Go through the folder containing the CSV files to be converted
    for f in os.listdir(csv_folder):
        if f.endswith(".csv"):
            # Make a sheet in the workbook with the same name as the CSV file
            ws = wb.create_sheet(f[:-4])

            # Open the current CSV file and read it into the CSV reader object
            with open(csv_folder + "/" + f) as filename:
                reader = csv.reader(filename)
                

                # Create a row index since each row can be iterated through to get all the columns
                i = 0
                for row in reader:
                    # Go through each column in each row (j: column index | each: value in a given column)
                    for j, each in enumerate(row):
                        v = each

                        # In case numerical data is being stored, attempt to convert the CSV text string to a float
                        # This makes Excel format the cell correctly
                        # If the conversion fails, the data is left as a string
                        try:
                            v = float(each)
                        except:
                            pass

                        # Write the data into the cell
                        ws.cell(row = i + 1, column = j + 1, value = v)

                    # Increment the row index after all columns in the row are written to the workbook
                    i += 1
        
            # Autofit column rows
            for column_cells in ws.columns:
                # The variable new_column_width stores the width necessary for the longest element in a column
                new_column_width = 0
                for cell in column_cells:
                    # Convert the value in a cell to a string to be able to determine its displayed width
                    v = as_text(cell.value)
                    
                    # Check if the current cell required width is greater than the previous maximum
                    new_column_width = max(len(v), new_column_width)
                    
                # Get the letter of the current column (Excel columns are lettered) and set the width
                new_column_letter = (openpyxl.utils.get_column_letter(column_cells[0].column))
                if new_column_width > 0:
                    ws.column_dimensions[new_column_letter].width = new_column_width + 1

    # Check if the output directory exists and create it if it doesn't
    output_path_parts = xlsx_location.split("/")[:-1]
    for i in range(1, len(output_path_parts)):
        path_to_check = ""
        for j in range(i + 1):
            path_to_check += (output_path_parts[j] + "/")
        
        path_to_check = path_to_check[:-1]

        if not os.path.isdir(path_to_check):
            os.mkdir(path_to_check)

    # Save the workbook
    wb.save(xlsx_location)


def as_text(value):
    if value is None:
        return ""
    else:
        return str(value)


def apply_font_to_cells(ws, font, cells = []):
    for cell in cells:
        c = ws[cell]
        c.font = font


def apply_font_to_cell_range(ws, font, row = (1, 1), column = (1, 1)):
    for r in range(row[0], row[1] + 1):
        for c in range(column[0], column[1] + 1):
            c = ws.cell(row = r, column = c)

            c.font = font

def apply_border_to_cells(ws, border, cells = []):
    for cell in cells:
        c = ws[cell]
        c.border = border


def apply_border_to_cell_range(ws, border, row = (1, 1), column = (1, 1)):
    for r in range(row[0], row[1] + 1):
        for c in range(column[0], column[1] + 1):
            c = ws.cell(row = r, column = c)

            c.border = border


def apply_alignment_to_cells(ws, alignment, cells = []):
    for cell in cells:
        c = ws[cell]
        c.alignment = alignment


def apply_alignment_to_cell_range(ws, alignment, row = (1, 1), column = (1, 1)):
    for r in range(row[0], row[1] + 1):
        for c in range(column[0], column[1] + 1):
            c = ws.cell(row = r, column = c)

            c.alignment = alignment


def main():
    # Create the argument parser
    parser = argparse.ArgumentParser()
    parser.add_argument("-n", "--nogui", help = "Invokes the program without a GUI (must supply argument --csvfolder)", action = "store_true")
    parser.add_argument("-c", "--csvfolder", help = "The location to the folder containing CSV files to be converted (output is placed into this folder by default)")
    parser.add_argument("-o", "--output", help = "Optional: The location of the output XLSX file (must contain the filename with .xlsx ending)")
    args = parser.parse_args()

    # Convert Windows paths to regular paths with forward slashes
    if args.csvfolder != None:
        args.csvfolder = args.csvfolder.replace("\\", "/")
    if args.output != None:
        args.output = args.output.replace("\\", "/")

    # Check whether the GUI should be used
    if args.nogui:
        if args.csvfolder == None:
            raise AttributeError("Option -c (--csvfolder) must be supplied if -n (--nogui) option is used")

        # Create a default output name
        csv_folder_name = args.csvfolder.split("/")[-1]
        out = args.csvfolder + "/" + csv_folder_name + ".xlsx"

        # If an output parameter was specified, use that instead
        if args.output != None:
            out = args.output
        
        # Call the conversion function
        try:
            convert_csv_files_to_xlsx(out, args.csvfolder)
            print("Successfully created XLSX at " + out)
        except:
            print("Failed to create XLSX")
    else:
        print("GUI not implemented yet. Please use -n (--nogui) option and supply -c (--csvfolder) argument.")


if __name__ == "__main__":
    main()
