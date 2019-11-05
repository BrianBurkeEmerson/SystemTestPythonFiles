#%%
import paramiko, time, re
import pandas as pd
import numpy as np
from datetime import datetime
import matplotlib.pyplot as plt
from collections import Counter


###### Login to gateway SSH######
def gwLogin(gateway,username,password):
    try:
        ssh.connect(gateway, username=username, password=password)
        time.sleep(1)
        print('Successfully connected to %s' % gw_name + '...')
    except paramiko.ssh_exception.NoValidConnectionsError:
        time.sleep(3)
        gwLogin(gateway, username, password)
    except TimeoutError:
        time.sleep(3)
        gwLogin(gateway, username, password)


###### Send command through SSH terminal ######
def sendCommand(cmd, chan, done):
    print('send command done', done)
    chan.send(cmd)
    print("command sent")
    chan.send("\n")
    time.sleep(.5)
    return done

###### Get results from command ######
def results(channel,ssh, done):
    print('results done', done)
    print('sending results')
    resp = channel.recv(9999)
    resp = resp.decode('ascii').split('\r\n')
    # print('resp',resp)
    parsed = parse(resp, done)
    return parsed

###### Parse results and determine what to do next ######
def parse(resp, done):
    done = list(done)
    for line in resp:
        if len(line) > 0:
            try:
                print(line)
                lineSplit = line.split()
                lineDate = lineSplit[0]
                lineTime = lineSplit[1]
                lineDateTime = lineDate + " " + lineTime
                lineDateTime = datetime.strptime(lineDateTime, '%Y-%m-%d %H:%M:%S')
                if "Notify device join MAC=0000:0000:0A10:00A0" in line:
                    mac = lineSplit[-1][-19:]
                    print('first mac', mac)
                    macList.append(mac)
                    startTime = lineDateTime
                    timeList.append(startTime)
                    print('start time', startTime)
                elif "0000:0000:0A10:00A0" in macList and '0000:0000:0000:0005' not in line:
                    if ('SMO device join MAC' in line) and ('Status=SecJoinReqReceived(4)' in line):
                        print('attempting to join')
                        mac = lineSplit[8]
                        if mac not in macList:
                            dictInit(mac,str(lineDateTime - timeList[0]))
                    elif ('SMO device join MAC' in line) and ('Status=Registered(20)' in line):
                        print('joined')
                        mac = lineSplit[8][-19:]
                        joinTime = str(lineDateTime - timeList[0])
                        timeList.append(joinTime)
                        if mac in macList:
                            if mac not in joinedList:
                                dictJoin(mac, joinTime)
                                printOut(mac)
                                done = checkNumDevs(done)
                    elif ('SMO device join failed MAC' in line):
                        print('failed join')
                        mac = lineSplit[-1]
                        print(mac)
                        if mac in macList:
                            dictCountInc(mac)
                    # print(deviceDict)
                    
            except IndexError:
                pass
            except ValueError:
                pass
                
            if "LOG File exceeded max size" in line:
                print('log full')
                done[1] = 'log full'
                return done
            elif "The system is going down" in line:
                done[1] = 'restarting'
                return done
            elif "tail: cannot open" in line:
                done[1] = "no log"
                print("log not there")
                return done
            elif "Last login" in line or 'cd /' in line or 'tail' in line:
                pass
                
    return done

###### Create a dictionary entry for mac address ######
def dictInit(macAddr,FirstJime):
    deviceDict[macAddr] = {}
    deviceDict[macAddr]['FirstJoinTime'] = FirstJime
    deviceDict[macAddr]['count'] = 0
    macList.append(macAddr)
    return deviceDict

###### Add join time to dictionary element ######
def dictJoin(macAddr, jTime):
    deviceDict[macAddr]['time'] = str(jTime)
    joinedList.append(macAddr)
    for key in deviceDict:
        if key == macAddr:
            print(key, macAddr)
            deviceDict[key]['count'] += 1
    print(deviceDict[macAddr])
    return deviceDict


###### Increment the device join count by 1 ######
def dictCountInc(macAddr):
    for key in deviceDict:
        if key == macAddr:
            deviceDict[key]['count'] += 1
    print(deviceDict[macAddr])
    return deviceDict


###### Re-Login to the gateway and send command ######
def restart(ssh, done):
    ssh.close()
    done = list(done)
    print('restart done', done)
    gwLogin(gw_name, username, password)
    return done


###### Determine if all devices have joined ######
def checkNumDevs(done):
    print("total devs = ", len(deviceDict))
    print("joined devs = ", len(joinedList))
    if len(joinedList)>= expectedDevs:
        done = list(done)
        done[0]='yes'
        print(done)
    return done


###### print dictionary to file ######
def printOut(macAddr):
    print("MAC:", macAddr, file=open(outFile, "a"))
    for key in deviceDict[macAddr]:
        print("\t", str(key) + ": " + str(deviceDict[macAddr][key]), file=open(outFile, "a"))
    print('\n', file=open(outFile, "a"))
        

#############################################################################################################################################
############################################################### PROGRAM START ###############################################################
#############################################################################################################################################

###### Initial login and send command ######
gw_name = '10.224.42.86'
password = 'root'
username = 'root'
ssh = paramiko.SSHClient()
ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

gwLogin(gw_name, username, password)
chan = ssh.invoke_shell()

done = 'no','running'
expectedDevs = 20

###### Follows mh.log in real time - filtered ######
command1 = 'tail -f -n 100 mh.log | egrep "truncated|reboot|join failed|device join"'

###### Cntrl + C - break current command ######
command2 = '\x03'

sendCommand('cd /access_node/tmp',chan, done)
sendCommand(command1,chan, done)

###### Create blank output file ######
outFile = r"C:\Users\E1256881\Desktop\ISA100Testing\Dan Files\2019 - 05 ISA 100\Python Scripts\isaJoin.txt"
print("",file=open(outFile, "w"))

###### Initialize arrays and dictionary ######
macList = []
timeList = []
joinedList =[]
deviceDict = {}

###### While loop - continue to read log until conditions met ######
while done[0] == 'no':
    done = results(chan,ssh,done)
    print('Status: ', done[0],"," ,done[1])

    if done[1] == 'log full':
        print('log full')
        sendCommand(command2, chan, done)
        sendCommand(command1, chan, done)
        done = list(done)
        done[1] = 'running'
        # restart(ssh)
    elif done[1] == 'restarting':
        print('restarting')
        deviceDict.clear()
        #time.sleep(45)
        done = list(done)
        done[1] = 'waiting'
        print('elif restarting', done)
        restart(ssh, done)
        try:
            chan = ssh.invoke_shell()
            sendCommand('cd /access_node/tmp', chan, done)
            sendCommand(command1, chan, done)
            done[1] = 'running'
            print('tried to send command')
        except OSError:
            done[1] = 'waiting'
        print('try done', done)
    elif done[1] == 'waiting':
        #time.sleep(5)
        restart(ssh, done)
    elif done[1] == 'no log':
        print('log not created yet')
        sendCommand(command1, chan, done)
        done = list(done)
        done[1] = 'running'
    elif done[1] == 'running':
        pass
    checkNumDevs(done)
    if done[1] == 'yes':
        exit()

#%%
##GRAPH ATTEMPS

missedList = []
for mac in deviceDict:
    missedList.append(deviceDict[mac]['count'])
missedList.sort()
print(missedList)
fig, ax = plt.subplots(figsize=(15, 5))

D = Counter(missedList)
plt.bar(range(len(D)), list(D.values()), align='center')
plt.xticks(range(len(D)), list(D.keys()))
plt.title('Attempts')
plt.ylabel('Number of Devices')
plt.xlabel('Number of Attempts')
rects = ax.patches
labels = ["label%d" % i for i in range(len(rects))]
for rect, label in zip(rects, labels):
    height = rect.get_height()
    ax.text(rect.get_x() + rect.get_width() / 2, height, height,
            ha='center', va='bottom')
plt.show()
plt.show()


#%%
for mac in deviceDict:
    print("MAC: " + mac + "---- Data ----" + str(deviceDict[mac]))
FirstJoinSec = []
PubTime = []
for mac in deviceDict:
    FirstJoinSec.append(int(deviceDict[mac]['FirstJoinTime'].split(":")[1])*60 + int(deviceDict[mac]['FirstJoinTime'].split(":")[2]))
    PubTime.append(int(deviceDict[mac]['time'].split(":")[1])*60 + int(deviceDict[mac]['time'].split(":")[2]))

#%%
missedList = []
for mac in deviceDict:
    missedList.append(deviceDict[mac]['time'])
missedList = sorted(missedList)
minlist = []
for i in missedList:
    minlist.append(int(i.split(":")[1]))
x = [0,5,15,30,45]
fig, ax = plt.subplots(figsize=(15, 5))
plt.title("Time to Join")
plt.ylabel('Number of Devices')
plt.xlabel('Time to Join (minutes)')
plt.hist(minlist,width = 2)
plt.xticks(np.arange(min(x), max(x)+1, 1.0))

rects = ax.patches
labels = ["label%d" % i for i in range(len(rects))]
for rect, label in zip(rects, labels):
    height = rect.get_height()
    if height != 0:
        ax.text(rect.get_x() + rect.get_width() / 2, height, height,
                ha='center', va='bottom')
plt.show()


#%%

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
wb = openpyxl.Workbook()
wb.save(r"C:\Users\E1256881\Desktop\ISA100Testing\Dan Files\2019 - 05 ISA 100\Python Scripts\ISA_DATA.xlsx")
ws1=wb.create_sheet('Sheet1')
df = pd.DataFrame.from_dict(deviceDict, orient='index') # convert dict to dataframe
rows = dataframe_to_rows(df)

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
         ws1.cell(row=r_idx, column=c_idx, value=value)

wb.save(r"C:\Users\E1256881\Desktop\ISA100Testing\Dan Files\2019 - 05 ISA 100\Python Scripts\ISA_DATA.xlsx")








        


    

